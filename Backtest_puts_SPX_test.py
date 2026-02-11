# %%
"""
Backtest  — Rolling Vanilla Puts (USD Budget Version)
================================================
On lance 252 backtests decales de 1 jour ouvre, chacun de meme duree.

Pour chaque couple (moneyness, maturity), on simule une strategie
de couverture systematique avec rebalancement a chaque echeance.

Methode:
- Budget annuel fixe en USD (ex: 1 000 000 $) identique pour tous les instruments
- MV_Equity dynamique: fixee a MV_EQUITY_T0 a la date t0, puis evolue avec le spot
- A chaque date de roll (tous les T jours), on achete des puts pour budget_roll / cout_unitaire
- A l'echeance, le put est exerce/expire et on realise le PnL = payoff - prime
- Le gain en capital economique M = Net Capital Benefit = economie SCR - prime payee

"""

# %%
import pandas as pd
import numpy as np
from scipy.stats import norm
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.cm as cm
import warnings
warnings.filterwarnings('ignore')
import os
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.colors import LinearSegmentedColormap, Normalize
import time

# %%
# ═══════════════════════════════════════════════════════════════════════
# 1. CHARGEMENT DES DONNEES
# ═══════════════════════════════════════════════════════════════════════

print("=" * 70)
print("1. CHARGEMENT DES DONNEES")
print("=" * 70)

df_raw = pd.read_csv(
    "C:\\Users\\jules\\Desktop\\Hedging Framework - Thesis\\Notebook\\Input\\SPX_IV_Cleaned.csv",
    sep=';', decimal=',', encoding='utf-8-sig'
)
df_raw['Date'] = pd.to_datetime(df_raw['Date'], format='%d/%m/%Y')
df_raw.rename(columns={
    'sigma': 'iv', 'dividend yield': 'q', 'risk-free rates': 'r'
}, inplace=True)
df_raw['iv'] /= 100
df_raw['q'] /= 100
df_raw['r'] /= 100
df_raw['tau'] = df_raw['maturity'] / 365

# %%
# Define Index
dates_df = df_raw[['Date', 'spot']].drop_duplicates('Date').sort_values('Date').reset_index(drop=True)
spot_s = dates_df.set_index('Date')['spot']

# %%
# ═══════════════════════════════════════════════════════════════════════
# SCR SOLVA II — Symmetric Adjustment
# ═══════════════════════════════════════════════════════════════════════

sa_raw = 0.5 * (spot_s - spot_s.rolling(756, min_periods=756).mean()) / spot_s.rolling(756, min_periods=756).mean()
sa_clipped = sa_raw.clip(-0.10, 0.10).dropna()
sa_map = sa_clipped.to_dict()

valid_dates_all = set(sa_clipped.index)
df_raw = df_raw[df_raw['Date'].isin(valid_dates_all)].copy()
df_raw['SA'] = df_raw['Date'].map(sa_map)
df_raw['equity_shock'] = 0.39 + df_raw['SA']

# IV moyenne globale (pour stress factor des couts de transaction)
iv_avg_global = df_raw['iv'].mean()

# All available trading dates (sorted)
all_trading_dates = sorted(df_raw['Date'].unique())

print(f"Donnees disponibles: {all_trading_dates[0].date()} -> {all_trading_dates[-1].date()}")
print(f"Jours de trading: {len(all_trading_dates)}")
print(f"Instruments: {len(df_raw.groupby(['moneyness', 'maturity']))} combinaisons")

# %%
# ═══════════════════════════════════════════════════════════════════════
# 2. PARAMETRES ROLLING ENTRY
# ═══════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("2. CONFIGURATION ROLLING ENTRY")
print("=" * 70)

N_STARTS = 252

# --- BUDGET USD FIXE ---
ANNUAL_BUDGET_USD = 1_000_000       # Budget annuel fixe en USD
MV_EQUITY_T0      = 100_000_000     # MV Equity initiale a t0 (100M USD)

# --- DATES MANUELLES ---
START_DATE = '2006-06-30'   # Format 'YYYY-MM-DD'
END_DATE   = '2025-12-30'   # Format 'YYYY-MM-DD'

start_dt = pd.to_datetime(START_DATE)
end_dt   = pd.to_datetime(END_DATE)

# Filtrer les dates de trading dans la plage
trading_dates_in_range = [d for d in all_trading_dates if start_dt <= d <= end_dt]
n_td = len(trading_dates_in_range)

# WINDOW en jours calendaires
WINDOW_DAYS = int((n_td - N_STARTS) / 252 * 365.25)
WINDOW_YEARS = WINDOW_DAYS / 365.25

print(f"Plage manuelle: {START_DATE} -> {END_DATE}")
print(f"Jours de trading dans la plage: {n_td}")
print(f"Window calculee: {WINDOW_YEARS:.2f} ans ({WINDOW_DAYS} jours calendaires)")

# Les 252 premiers jours de trading de la plage sont les starts
candidate_starts = trading_dates_in_range[:N_STARTS]

# Verification que chaque start + WINDOW tient dans la plage
last_end = candidate_starts[-1] + pd.Timedelta(days=WINDOW_DAYS)
if last_end > end_dt:
    WINDOW_DAYS = (end_dt - candidate_starts[-1]).days
    WINDOW_YEARS = WINDOW_DAYS / 365.25
    print(f"Window ajustee: {WINDOW_YEARS:.2f} ans (contrainte par END_DATE)")

N_STARTS_ACTUAL = len(candidate_starts)

# Spot de reference a t0 (premier jour de trading de la plage)
SPOT_T0 = spot_s.loc[trading_dates_in_range[0]]

print(f"\nWindow: {WINDOW_YEARS:.2f} ans ({WINDOW_DAYS} jours)")
print(f"Starts: {N_STARTS_ACTUAL}")
print(f"Premier start: {candidate_starts[0].date()}")
print(f"Dernier start: {candidate_starts[-1].date()}")
print(f"Fin du premier backtest: {(candidate_starts[0] + pd.Timedelta(days=WINDOW_DAYS)).date()}")
print(f"Fin du dernier backtest: {(candidate_starts[-1] + pd.Timedelta(days=WINDOW_DAYS)).date()}")
print(f"Budget annuel: ${ANNUAL_BUDGET_USD:,.0f}")
print(f"MV Equity t0: ${MV_EQUITY_T0:,.0f}")
print(f"Spot t0: {SPOT_T0:.2f}")
print(f"Budget annuel / MV_t0: {ANNUAL_BUDGET_USD / MV_EQUITY_T0 * 100:.2f}%")

folder_name = f"Output_{START_DATE}_{END_DATE}"
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

# %%
# ═══════════════════════════════════════════════════════════════════════
# 3. FONCTIONS
# ═══════════════════════════════════════════════════════════════════════

def bs_put_price(S, K, tau, r, q, sigma):
    """Prix Black-Scholes d'un put europeen."""
    tau = np.maximum(tau, 1e-10)
    d1 = (np.log(S / K) + (r - q + 0.5 * sigma**2) * tau) / (sigma * np.sqrt(tau))
    d2 = d1 - sigma * np.sqrt(tau)
    return K * np.exp(-r * tau) * norm.cdf(-d2) - S * np.exp(-q * tau) * norm.cdf(-d1)


def put_intrinsic(S, K):
    """Payoff intrinseque du put a l'echeance."""
    return np.maximum(K - S, 0)


def apply_transaction_costs_pro(premium_per_unit, moneyness, iv_current, iv_mean, nb_puts):
    """Couts de transaction."""
    unit_brokerage = 0.000
    total_brokerage = nb_puts * unit_brokerage

    '''if moneyness >= 95:
        base_spread_pct = 0.015
    elif moneyness >= 85:
        base_spread_pct = 0.04
    elif moneyness >= 70:
        base_spread_pct = 0.08
    else:
        base_spread_pct = 0.15

    stress_factor = 1.0
    if iv_current > (iv_mean * 1.2):
        stress_factor = 2.0
    elif iv_current > iv_mean:
        stress_factor = 1.5'''

    total_spread_cost = (premium_per_unit * nb_puts) ##* base_spread_pct * stress_factor
    return total_spread_cost ##+ total_brokerag


def compute_mv_equity(S_current, spot_t0, mv_equity_t0):
    """MV Equity dynamique: evolue proportionnellement au spot."""
    return mv_equity_t0 * (S_current / spot_t0)


def run_single_backtest(df_period, instruments_list, annual_budget_usd,
                        spot_t0, mv_equity_t0, iv_avg):
    """
    Execute un backtest rolling sur une periode donnee.
    Budget fixe en USD, MV dynamique, M = net capital benefit.
    Retourne la liste des records de rolls (detail complet).
    """
    results = []

    for _, instr in instruments_list.iterrows():
        money = instr['moneyness']
        mat = instr['maturity']

        sub = df_period[(df_period['moneyness'] == money) & (df_period['maturity'] == mat)].copy()
        sub = sub.sort_values('Date').reset_index(drop=True)
        n = len(sub)

        if n == 0:
            continue

        roll_freq_td = max(1, int(mat * 252 / 365))
        rolls_per_year = 365 / mat
        budget_per_roll_usd = annual_budget_usd / rolls_per_year

        i = 0
        while i < n:
            row_buy = sub.iloc[i]
            date_buy = row_buy['Date']
            S_buy = row_buy['spot']
            iv_buy = row_buy['iv']
            r_buy = row_buy['r']
            q_buy = row_buy['q']
            tau_buy = row_buy['tau']
            sa_buy = row_buy['SA']
            shock_buy = row_buy['equity_shock']

            # MV Equity dynamique a cette date
            MV_eq = compute_mv_equity(S_buy, spot_t0, mv_equity_t0)

            # CALCUL DU STRIKE VIA LE FORWARD
            F_buy = S_buy * np.exp((r_buy - q_buy) * tau_buy)
            K = (money / 100) * F_buy

            put_price = bs_put_price(S_buy, K, tau_buy, r_buy, q_buy, iv_buy)

            if put_price <= 0 or np.isnan(put_price):
                i += roll_freq_td
                continue

            # Cout unitaire all-in (prime + frais) pour 1 put
            unit_cost_1put = apply_transaction_costs_pro(
                premium_per_unit=put_price,
                moneyness=money,
                iv_current=iv_buy,
                iv_mean=iv_avg,
                nb_puts=1
            )

            # Budget USD fixe par roll -> nombre de puts
            budget_abs = budget_per_roll_usd
            nb_puts = budget_abs / unit_cost_1put

            premium_paid_usd = nb_puts * put_price
            total_fees_usd = nb_puts * unit_cost_1put - premium_paid_usd
            total_cost_usd = nb_puts * unit_cost_1put  # = budget_abs

            # ─── SCR & NET CAPITAL BENEFIT (M) ───
            # SCR brut (sans couverture)
            SCR_nu_usd = MV_eq * shock_buy

            # Valeur du put sous choc
            S_shocked = S_buy * (1 - shock_buy)
            put_shocked = bs_put_price(S_shocked, K, tau_buy, r_buy, q_buy, iv_buy)
            hedge_gain_per_put = put_shocked - put_price

            # Economie brute de SCR
            total_hedge_gain_usd = nb_puts * hedge_gain_per_put
            SCR_hedged_usd = max(0, SCR_nu_usd - total_hedge_gain_usd)
            gross_scr_saving_usd = SCR_nu_usd - SCR_hedged_usd

            # Net Capital Benefit = economie SCR - prime payee
            net_capital_benefit_usd = gross_scr_saving_usd - premium_paid_usd
            M_net_pct = net_capital_benefit_usd / MV_eq * 100

            # Efficiency ratio: combien de SCR economise par euro de prime
            efficiency_ratio = gross_scr_saving_usd / premium_paid_usd if premium_paid_usd > 0 else 0

            # ─── PAYOFF & PNL ───
            expiry_idx = min(i + roll_freq_td, n - 1)
            row_exp = sub.iloc[expiry_idx]
            date_exp = row_exp['Date']
            S_exp = row_exp['spot']

            days_passed = (date_exp - date_buy).days
            tau_remaining = max(0, (mat - days_passed) / 365)

            if tau_remaining > 1/365:
                payoff_per_put = bs_put_price(S_exp, K, tau_remaining, row_exp['r'], row_exp['q'], row_exp['iv'])
            else:
                payoff_per_put = put_intrinsic(S_exp, K)

            payoff_usd = nb_puts * payoff_per_put
            pnl_roll_usd = payoff_usd - budget_abs
            pnl_roll_pct = pnl_roll_usd / MV_eq * 100

            results.append({
                # ID & Dates
                'moneyness': money,
                'maturity': mat,
                'date_buy': date_buy,
                'date_exp': date_exp,
                'tau_remaining': tau_remaining,
                'is_mtm_exit': 1 if tau_remaining > 1/365 else 0,

                # Market levels
                'S_buy': S_buy,
                'S_exp': S_exp,
                'F_buy': F_buy,
                'K': K,
                'iv_buy': iv_buy,
                'r_buy': r_buy,
                'q_buy': q_buy,
                'return_underlying': (S_exp / S_buy - 1) * 100,

                # MV Equity
                'MV_equity': MV_eq,
                'MV_equity_pct_t0': MV_eq / mv_equity_t0 * 100,

                # Put pricing
                'put_price_per_unit': put_price,
                'put_price_pct_spot': put_price / S_buy * 100,
                'put_shocked_per_unit': put_shocked,
                'hedge_gain_per_put': hedge_gain_per_put,

                # Budget & Costs (USD)
                'budget_per_roll_usd': budget_abs,
                'nb_puts': nb_puts,
                'premium_paid_usd': premium_paid_usd,
                'transaction_fees_usd': total_fees_usd,
                'total_cost_usd': total_cost_usd,
                'premium_pct_mv': premium_paid_usd / MV_eq * 100,
                'fees_pct_premium': total_fees_usd / premium_paid_usd * 100 if premium_paid_usd > 0 else 0,

                # Solvabilite II (SCR) — en USD
                'equity_shock': shock_buy,
                'SCR_nu_usd': SCR_nu_usd,
                'SCR_hedged_usd': SCR_hedged_usd,
                'gross_scr_saving_usd': gross_scr_saving_usd,
                'net_capital_benefit_usd': net_capital_benefit_usd,

                # SCR — en % de MV
                'SCR_nu_pct': shock_buy * 100,
                'SCR_hedged_pct': SCR_hedged_usd / MV_eq * 100,
                'gross_scr_saving_pct': gross_scr_saving_usd / MV_eq * 100,
                'M_net_pct': M_net_pct,
                'efficiency_ratio': efficiency_ratio,

                # Performance — Payoff & PnL
                'payoff_per_put': payoff_per_put,
                'payoff_usd': payoff_usd,
                'pnl_roll_usd': pnl_roll_usd,
                'pnl_roll_pct': pnl_roll_pct,
            })

            i += roll_freq_td

    return results


# ═══════════════════════════════════════════════════════════════════════
# 4. BOUCLE PRINCIPALE — 252 BACKTESTS
# ═══════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("4. LANCEMENT DES 252 BACKTESTS")
print("=" * 70)

instruments = df_raw.groupby(['moneyness', 'maturity']).size().reset_index(name='count')

# Storage
all_agg_records = []
all_detail_records = []   # <-- NOUVEAU: tous les rolls detailles pour l'Excel
cum_pnl_series = {}

t_global = time.time()

for start_idx, start_date in enumerate(candidate_starts):
    end_date = start_date + pd.Timedelta(days=WINDOW_DAYS)

    # Spot de reference pour ce backtest = spot a la date de start
    spot_ref = spot_s.loc[start_date] if start_date in spot_s.index else SPOT_T0

    # Filter data for this window
    df_window = df_raw[(df_raw['Date'] >= start_date) & (df_raw['Date'] <= end_date)]

    if len(df_window) == 0:
        continue

    # Run backtest
    rolls = run_single_backtest(
        df_window, instruments, ANNUAL_BUDGET_USD,
        spot_ref, MV_EQUITY_T0, iv_avg_global
    )

    if not rolls:
        continue

    bt_single = pd.DataFrame(rolls)
    window_years = (bt_single['date_exp'].max() - bt_single['date_buy'].min()).days / 365.25

    # Store detail records with start_id
    for rec in rolls:
        rec_copy = rec.copy()
        rec_copy['start_id'] = start_idx
        rec_copy['start_date_bt'] = start_date
        all_detail_records.append(rec_copy)

    # Aggregate per instrument for this start
    for (money, mat), grp in bt_single.groupby(['moneyness', 'maturity']):
        pnl_cum_pct = grp['pnl_roll_pct'].sum()
        pnl_ann_pct = pnl_cum_pct / window_years if window_years > 0 else 0
        pnl_cum_usd = grp['pnl_roll_usd'].sum()
        pnl_ann_usd = pnl_cum_usd / window_years if window_years > 0 else 0

        # Cumulative PnL series for drawdown calc
        cum_pnl = grp['pnl_roll_pct'].cumsum()
        running_max = cum_pnl.cummax()
        drawdowns = cum_pnl - running_max
        max_dd = drawdowns.min()

        # Store cumulative PnL time series
        cum_pnl_series[(money, mat, start_idx)] = (grp['date_exp'].values, cum_pnl.values)

        all_agg_records.append({
            'start_id': start_idx,
            'start_date': start_date,
            'end_date': end_date,
            'moneyness': money,
            'maturity': mat,
            'n_rolls': len(grp),
            # M net (net capital benefit)
            'M_mean': grp['M_net_pct'].mean(),
            'M_median': grp['M_net_pct'].median(),
            'M_std': grp['M_net_pct'].std(),
            # Gross SCR saving (for comparison)
            'gross_scr_saving_mean': grp['gross_scr_saving_pct'].mean(),
            # Efficiency ratio
            'efficiency_ratio_mean': grp['efficiency_ratio'].mean(),
            # PnL
            'pnl_cumulated_pct': pnl_cum_pct,
            'pnl_annualized_pct': pnl_ann_pct,
            'pnl_cumulated_usd': pnl_cum_usd,
            'pnl_annualized_usd': pnl_ann_usd,
            'pnl_mean_per_roll_pct': grp['pnl_roll_pct'].mean(),
            'pnl_std_per_roll_pct': grp['pnl_roll_pct'].std(),
            'pnl_mean_per_roll_usd': grp['pnl_roll_usd'].mean(),
            'hit_rate': (grp['pnl_roll_pct'] > 0).mean() * 100,
            'max_drawdown': max_dd,
            'max_single_gain_pct': grp['pnl_roll_pct'].max(),
            'max_single_loss_pct': grp['pnl_roll_pct'].min(),
            # SCR
            'SCR_nu_mean': grp['SCR_nu_pct'].mean(),
            'SCR_hedged_mean': grp['SCR_hedged_pct'].mean(),
            # Premium
            'avg_premium_pct_mv': grp['premium_pct_mv'].mean(),
            'total_premium_usd': grp['premium_paid_usd'].sum(),
            'avg_return': grp['return_underlying'].mean(),
        })

    # Progress
    if (start_idx + 1) % 25 == 0 or start_idx == 0:
        elapsed = time.time() - t_global
        eta = elapsed / (start_idx + 1) * (N_STARTS_ACTUAL - start_idx - 1)
        print(f"  Start {start_idx+1}/{N_STARTS_ACTUAL} ({start_date.date()}) — "
              f"{elapsed:.0f}s elapsed, ETA {eta:.0f}s")

total_time = time.time() - t_global
print(f"\nTermine: {N_STARTS_ACTUAL} backtests en {total_time:.0f}s ({total_time/60:.1f} min)")

# %%
# ═══════════════════════════════════════════════════════════════════════
# 5. ANALYSE DISTRIBUTIONNELLE
# ═══════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("5. ANALYSE DISTRIBUTIONNELLE")
print("=" * 70)

df_agg = pd.DataFrame(all_agg_records)
df_detail = pd.DataFrame(all_detail_records)

print(f"Total records aggreges: {len(df_agg):,}")
print(f"Total records detailles: {len(df_detail):,}")
print(f"Instruments: {df_agg.groupby(['moneyness','maturity']).ngroups}")
print(f"Starts effectifs par instrument: ~{df_agg.groupby(['moneyness','maturity'])['start_id'].count().median():.0f}")

# Distribution des metriques par instrument
dist = df_agg.groupby(['moneyness', 'maturity']).agg(
    n_starts=('start_id', 'count'),

    # PnL annualise (%)
    pnl_ann_mean=('pnl_annualized_pct', 'mean'),
    pnl_ann_median=('pnl_annualized_pct', 'median'),
    pnl_ann_std=('pnl_annualized_pct', 'std'),
    pnl_ann_p5=('pnl_annualized_pct', lambda x: np.percentile(x, 5)),
    pnl_ann_p25=('pnl_annualized_pct', lambda x: np.percentile(x, 25)),
    pnl_ann_p75=('pnl_annualized_pct', lambda x: np.percentile(x, 75)),
    pnl_ann_p95=('pnl_annualized_pct', lambda x: np.percentile(x, 95)),

    # PnL annualise (USD)
    pnl_ann_usd_mean=('pnl_annualized_usd', 'mean'),
    pnl_ann_usd_median=('pnl_annualized_usd', 'median'),

    # M net (Net Capital Benefit)
    M_mean_avg=('M_mean', 'mean'),
    M_mean_median=('M_mean', 'median'),
    M_mean_std=('M_mean', 'std'),
    M_mean_p5=('M_mean', lambda x: np.percentile(x, 5)),
    M_mean_p95=('M_mean', lambda x: np.percentile(x, 95)),

    # Gross SCR saving
    gross_scr_saving_avg=('gross_scr_saving_mean', 'mean'),

    # Efficiency ratio
    efficiency_ratio_avg=('efficiency_ratio_mean', 'mean'),

    # Hit rate
    hit_rate_avg=('hit_rate', 'mean'),
    hit_rate_std=('hit_rate', 'std'),

    # Max Drawdown
    max_dd_mean=('max_drawdown', 'mean'),
    max_dd_worst=('max_drawdown', 'min'),
    max_dd_p5=('max_drawdown', lambda x: np.percentile(x, 5)),

    # SCR
    SCR_hedged_avg=('SCR_hedged_mean', 'mean'),
    SCR_hedged_std=('SCR_hedged_mean', 'std'),
    SCR_nu_mean=('SCR_nu_mean', 'mean'),

    # Premium
    avg_premium_pct=('avg_premium_pct_mv', 'mean'),

).round(4)

print("\n── Distribution du PnL annualise ──")
print(dist[['n_starts', 'pnl_ann_mean', 'pnl_ann_median', 'pnl_ann_p5', 'pnl_ann_p95', 'pnl_ann_std']].to_string())

print("\n── Distribution du M net (Net Capital Benefit) ──")
print(dist[['M_mean_avg', 'M_mean_median', 'M_mean_p5', 'M_mean_p95', 'M_mean_std']].to_string())

print("\n── Gross SCR Saving & Efficiency Ratio ──")
print(dist[['gross_scr_saving_avg', 'efficiency_ratio_avg', 'avg_premium_pct']].to_string())

print("\n── Risque: Max Drawdown ──")
print(dist[['max_dd_mean', 'max_dd_worst', 'max_dd_p5']].to_string())

# %%
# ═══════════════════════════════════════════════════════════════════════
# 6. VISUALISATIONS
# ═══════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("6. VISUALISATIONS")
print("=" * 70)

COL = {
    'median':   '#2166AC',
    'mean':     '#1B9E77',
    'ci_band':  '#92C5DE',
    'iqr_band': '#4393C3',
    'pnl_pos':  '#1B9E77',
    'pnl_neg':  '#D95F02',
    'dd':       '#C0392B',
    'ref':      '#B2182B',
}

focus_moneys = [50, 60, 70, 80, 90, 100]

ANNUAL_BUDGET_PCT_DISPLAY = ANNUAL_BUDGET_USD / MV_EQUITY_T0 * 100  # for chart labels

# %%
# ── Frontiere d'Efficacite "Blue Selection" ──
fig0, ax = plt.subplots(figsize=(13, 8.5))

blue_shades = ['#E1F5FE', "#4DA683", "#0B755B", "#A6C9DC", "#70A5CE", '#002544']
moneyness_list = sorted(dist.index.get_level_values(0).unique())
colors_money = dict(zip(moneyness_list, blue_shades))

dist_filtered = dist[dist['pnl_ann_median'] > 0].copy()
dist_filtered['efficiency_score'] = dist_filtered['M_mean_avg'] * dist_filtered['pnl_ann_median'] / 100

for (money, mat), row in dist_filtered.iterrows():
    color = colors_money.get(money, '#0288D1')
    score = row['efficiency_score']
    point_size = (score * 800)

    ax.scatter(row['M_mean_avg'], row['pnl_ann_median'],
               color=color, marker='o', s=max(point_size, 30),
               edgecolors='black', linewidth=0.6, alpha=0.85, zorder=5)

    ax.annotate(f"{mat}j",
                (row['M_mean_avg'], row['pnl_ann_median']),
                fontsize=8, ha='center', va='center',
                fontweight='normal', color='white', zorder=7)

    ax.annotate(f"sc:{score:.1f}",
                (row['M_mean_avg'], row['pnl_ann_median']),
                xytext=(0, -8), textcoords='offset points',
                fontsize=9, ha='center', va='top',
                fontweight='normal', color='black', alpha=0.7, zorder=6)

ax.axhline(y=0, color='black', linewidth=0.8)
ax.grid(True, linestyle=':', alpha=0.4)
ax.set_xlabel('Average Net Capital Benefit M (%)', fontsize=12, fontweight='500')
ax.set_ylabel('Annualized PnL Median (%)', fontsize=12, fontweight='500')
ax.set_title(f'Strategy Efficiency Frontier (PnL > 0)\n'
             f'M = Net Capital Benefit (SCR saving - premium) | Budget = ${ANNUAL_BUDGET_USD/1e6:.0f}M/yr',
             fontsize=14, fontweight='bold', pad=20)

for m in moneyness_list:
    ax.scatter([], [], color=colors_money[m], s=100, label=f'K={m}%',
               edgecolors='black', linewidth=0.5)
ax.legend(loc='upper left', fontsize=9, title='Moneyness', title_fontsize=10, ncol=2, frameon=True)

plt.tight_layout()
plt.savefig(f'{folder_name}/efficient_frontier_final_blue.png', dpi=150, bbox_inches='tight')
plt.show()
print("Graphique genere : efficient_frontier_final_blue.png")

# %%
# ── Plot 1: Frontiere M vs PnL ──
fig1, ax = plt.subplots(figsize=(14, 9))

colors_money = {50: '#1B2A4A', 60: '#D95F02', 70: '#1B9E77', 80: '#E7298A', 90: '#7570B3', 100: '#E6AB02'}
markers_mat = {30: 'o', 60: 's', 90: '^', 120: 'D', 150: 'v', 180: 'p', 360: '*', 540: 'h', 720: 'X'}

for (money, mat), row in dist.iterrows():
    color = colors_money.get(money, 'gray')
    marker = markers_mat.get(mat, 'o')

    ax.scatter(row['M_mean_avg'], row['pnl_ann_median'],
               color=color, marker=marker, s=120,
               edgecolors='black', linewidth=0.5, alpha=0.9, zorder=5)

    ax.plot([row['M_mean_avg'], row['M_mean_avg']],
            [row['pnl_ann_p5'], row['pnl_ann_p95']],
            color=color, linewidth=1.5, alpha=0.4, zorder=3)

    ax.plot([row['M_mean_p5'], row['M_mean_p95']],
            [row['pnl_ann_median'], row['pnl_ann_median']],
            color=color, linewidth=1.5, alpha=0.4, zorder=3)

    ax.annotate(f'{mat}j', (row['M_mean_avg'], row['pnl_ann_median']),
                fontsize=5, ha='center', va='bottom',
                textcoords='offset points', xytext=(0, 6))

ax.axhline(y=0, color='black', linewidth=0.8)
ax.axhline(y=-ANNUAL_BUDGET_PCT_DISPLAY, color='gray', linewidth=0.5, linestyle='--', alpha=0.5)

for money, col in colors_money.items():
    ax.scatter([], [], color=col, s=80, label=f'K={money}%', edgecolors='black', linewidth=0.5)
ax.legend(loc='upper left', fontsize=8, title='Moneyness', title_fontsize=9, ncol=2)

ax.set_xlabel('Average Net Capital Benefit M (%)', fontsize=12)
ax.set_ylabel('PnL annualized — median (%, CI 90% in bars)', fontsize=12)
ax.set_title(f'Efficient Frontier — {START_DATE} - {END_DATE}\n'
             f'Budget = ${ANNUAL_BUDGET_USD/1e6:.0f}M/year | MV_Eq(t0) = ${MV_EQUITY_T0/1e6:.0f}M',
             fontsize=13, fontweight='bold')
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f'{folder_name}/efficient_frontier.png', dpi=150, bbox_inches='tight')
print("Plot 1: efficient_frontier.png")

# %%
# ── Plot 2: Surface 3D ──
pivot_M = dist['M_mean_avg'].unstack()
pivot_pnl = dist['pnl_ann_median'].unstack()

X_mats = pivot_M.columns.values
Y_money = pivot_M.index.values
X, Y = np.meshgrid(X_mats, Y_money)
Z = pivot_M.values
C = pivot_pnl.values

colors_list = ["#E0E0E0", "#90CAF9", "#1E88E5", "#0000FF"]
custom_blue_map = LinearSegmentedColormap.from_list("VibrantBlue", colors_list)

pnl_positive = C[C > 0]
if len(pnl_positive) > 0:
    vmax_saturated = np.percentile(pnl_positive, 20)
else:
    vmax_saturated = np.nanmax(C)

norm_c = Normalize(vmin=0, vmax=vmax_saturated, clip=True)
colors_surface = custom_blue_map(norm_c(C))

fig2 = plt.figure(figsize=(16, 11))
ax = fig2.add_subplot(111, projection='3d')
surf = ax.plot_surface(X, Y, Z, facecolors=colors_surface, shade=True,
                       edgecolor='white', linewidth=0.03, alpha=0.95, antialiased=True)

mappable = cm.ScalarMappable(cmap=custom_blue_map, norm=norm_c)
mappable.set_array(C)
cbar = fig2.colorbar(mappable, ax=ax, shrink=0.4, aspect=15, pad=0.1)
cbar.set_label(f'PnL Annualized Median (Saturated at {vmax_saturated:.1f}%)', fontweight='bold')

ax.view_init(elev=25, azim=-45)
ax.set_xlabel('\nMaturity (days)', fontsize=10)
ax.set_ylabel('\nMoneyness (%)', fontsize=10)
ax.set_zlabel('Average Net Capital Benefit M (%)', fontsize=10)
ax.set_title('High-Performance 4D Frontier\nM = Net Capital Benefit', fontsize=15, fontweight='bold', pad=30)

ax.xaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
ax.yaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
ax.zaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))

plt.tight_layout()
plt.savefig(f"{folder_name}/surface_4D_saturated_blue.png", dpi=200, bbox_inches='tight')
plt.show()
print("Graphique 4D genere")

# %%
# ── Plot 3: Heatmaps croisees ──
fig3, (ax_scr, ax_pnl) = plt.subplots(1, 2, figsize=(20, 8))

pivot_M_h = dist['M_mean_avg'].unstack()
pivot_pnl_h = dist['pnl_ann_median'].unstack()

colors_ = ["#E0E0E0", "#90CAF9", "#1E88E5", "#0D47A1"]
cmap = LinearSegmentedColormap.from_list("GrayToBlue", colors_)

im1 = ax_scr.imshow(pivot_M_h.values, cmap=cmap, aspect='auto')
ax_scr.set_title('Average Net Capital Benefit M (%)', fontsize=14, fontweight='bold', pad=15)

vmax_pnl = np.nanpercentile(pivot_pnl_h.values[pivot_pnl_h.values > 0], 90) if np.any(pivot_pnl_h.values > 0) else 1
im2 = ax_pnl.imshow(pivot_pnl_h.values, cmap=cmap, aspect='auto', vmin=0, vmax=vmax_pnl)
ax_pnl.set_title('PnL Annualized Median (%)', fontsize=14, fontweight='bold', pad=15)

for ax, im, data, fmt in zip([ax_scr, ax_pnl], [im1, im2], [pivot_M_h, pivot_pnl_h], ['.2f', '.2f']):
    ax.set_xticks(range(len(data.columns)))
    ax.set_xticklabels([f'{int(c)}j' for c in data.columns], fontsize=9)
    ax.set_yticks(range(len(data.index)))
    ax.set_yticklabels([f'K={int(m)}%' for m in data.index], fontsize=9)
    ax.set_xlabel('Maturity')
    ax.set_ylabel('Moneyness')

    thresh = np.nanmean(data.values)
    for i in range(len(data.index)):
        for j in range(len(data.columns)):
            val = data.values[i, j]
            if not np.isnan(val):
                color = "white" if val > thresh else "black"
                ax.text(j, i, format(val, fmt), ha='center', va='center',
                        fontsize=9, fontweight='bold', color=color)

fig3.colorbar(im1, ax=ax_scr, label='M net (%)', shrink=0.8)
fig3.colorbar(im2, ax=ax_pnl, label='PnL (%)', shrink=0.8)

fig3.suptitle(f'Risk-Return Matrix: Net Capital Benefit vs Profitability\n'
             f'Budget = ${ANNUAL_BUDGET_USD/1e6:.0f}M/year | Period: {START_DATE} - {END_DATE}',
             fontsize=16, fontweight='bold', y=1.02)

plt.tight_layout()
plt.savefig(f"{folder_name}/heatmaps_scr_pnl_summary.png", dpi=150, bbox_inches='tight')
plt.show()
print("Double Heatmap generee")

# %%
# ── Plot 4: Drawdowns S&P 500 ──
peak = spot_s.expanding(min_periods=1).max()
drawdown_series = (spot_s / peak) - 1

dd_threshold = -0.10
is_in_dd = drawdown_series <= dd_threshold
diff = is_in_dd.astype(int).diff()
dd_starts = diff[diff == 1].index
dd_ends = diff[diff == -1].index
if is_in_dd.iloc[-1]:
    dd_ends = dd_ends.append(pd.Index([is_in_dd.index[-1]]))

fig4, ax1 = plt.subplots(figsize=(14, 7))
ax1.plot(spot_s.index, spot_s.values, color='#1B2A4A', linewidth=1.5, label='S&P 500 Spot')
ax1.set_ylabel('Spot S&P 500 ($)', color='#1B2A4A', fontweight='bold', fontsize=11)
ax1.tick_params(axis='y', labelcolor='#1B2A4A')
ax1.grid(True, alpha=0.2, linestyle='--')

ax2 = ax1.twinx()
ax2.plot(drawdown_series.index, drawdown_series.values * 100, color='#C0392B',
         linewidth=1, linestyle='-', alpha=0.4, label='Drawdown (%)')
ax2.fill_between(drawdown_series.index, drawdown_series.values * 100, 0,
                 where=(drawdown_series <= 0), color='#C0392B', alpha=0.1)
ax2.set_ylabel('Drawdown (%)', color='#C0392B', fontweight='bold', fontsize=11)
ax2.tick_params(axis='y', labelcolor='#C0392B')
ax2.set_ylim(-60, 5)

for i, (s, e) in enumerate(zip(dd_starts, dd_ends)):
    ax1.axvspan(s, e, color='gray', alpha=0.15,
                label='Stress Zone (DD > 10%)' if i == 0 else "")

plt.title(f'Market Regime: S&P 500 Spots and Drawdowns\nThreshold: {abs(dd_threshold)*100:.0f}%',
          fontsize=14, fontweight='bold', pad=15)
ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
ax1.xaxis.set_major_locator(mdates.YearLocator())

lines, labels = ax1.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax1.legend(lines + lines2, labels + labels2, loc='upper left', frameon=True, fontsize=10)

plt.tight_layout()
plt.savefig(f"{folder_name}/market_drawdowns_analysis.png", dpi=150, bbox_inches='tight')
plt.show()
print("Graphique de marche genere")

# %%
# ── Plot 5: PnL cumule avec IC vs Drawdowns ──
focus_mats = [90, 120, 150, 180]

COL_median = '#2166AC'
COL_ci     = '#92C5DE'
COL_iqr    = '#4393C3'
COL_dd     = '#C0392B'

fig5, axes_big = plt.subplots(len(focus_moneys), len(focus_mats),
                              figsize=(28, 30), sharex=True)
fig5.suptitle(
    f'Cumulative PnL with CI (252 entries) vs S&P 500 Drawdowns\n'
    f'Budget = ${ANNUAL_BUDGET_USD/1e6:.0f}M/year | Window = {WINDOW_YEARS:.1f}y',
    fontsize=18, fontweight='bold')

for row_idx, money in enumerate(focus_moneys):
    for col_idx, mat in enumerate(focus_mats):
        ax = axes_big[row_idx, col_idx]

        ax_dd = ax.twinx()
        dd_plot = drawdown_series.loc[
            candidate_starts[0]:candidate_starts[-1] + pd.Timedelta(days=WINDOW_DAYS)
        ]
        ax_dd.fill_between(dd_plot.index, dd_plot.values * 100, 0,
                           color=COL_dd, alpha=0.06)
        ax_dd.set_ylim(-60, 5)
        if col_idx == len(focus_mats) - 1:
            ax_dd.set_ylabel('Drawdown (%)', color=COL_dd, fontsize=8)
            ax_dd.tick_params(axis='y', labelcolor=COL_dd, labelsize=7)
        else:
            ax_dd.set_yticklabels([])

        all_paths = []
        for sid in range(N_STARTS_ACTUAL):
            key = (money, mat, sid)
            if key in cum_pnl_series:
                dates, cum = cum_pnl_series[key]
                all_paths.append((dates, cum))

        if not all_paths:
            ax.set_title(f'K={money}% T={mat}j — No data', fontsize=9)
            continue

        all_dates_set = set()
        for dates, _ in all_paths:
            all_dates_set.update(dates)
        common_dates = np.array(sorted(all_dates_set))

        matrix = np.full((len(all_paths), len(common_dates)), np.nan)
        for i, (dates, cum) in enumerate(all_paths):
            path_series = pd.Series(cum, index=dates)
            path_reindexed = path_series.reindex(common_dates, method='ffill')
            matrix[i, :] = path_reindexed.values

        p5  = np.nanpercentile(matrix, 5, axis=0)
        p25 = np.nanpercentile(matrix, 25, axis=0)
        p50 = np.nanpercentile(matrix, 50, axis=0)
        p75 = np.nanpercentile(matrix, 75, axis=0)
        p95 = np.nanpercentile(matrix, 95, axis=0)

        ax.fill_between(common_dates, p5, p95, alpha=0.12, color=COL_ci, label='P5–P95')
        ax.fill_between(common_dates, p25, p75, alpha=0.25, color=COL_iqr, label='P25–P75')
        ax.plot(common_dates, p50, color=COL_median, linewidth=1.8, label='Median')

        ax.axhline(y=0, color='black', linewidth=0.6, alpha=0.5)
        ax.set_title(f'K={money}%  T={mat}j', fontsize=10, fontweight='bold')
        ax.grid(True, alpha=0.15)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))

        if col_idx == 0:
            ax.set_ylabel('Cumulative PnL (%)', fontsize=9)
        if row_idx == 0 and col_idx == 0:
            ax.legend(fontsize=7, loc='upper left', frameon=True)

plt.tight_layout(rect=[0, 0.02, 1, 0.96])
plt.savefig(f"{folder_name}/rolling_pnl_vs_drawdowns.png", dpi=150, bbox_inches='tight')
print(f"Graphique exporte: {folder_name}/rolling_pnl_vs_drawdowns.png")

# %%
# ── Plot 6: Boxplots PnL ──
fig6, axes2 = plt.subplots(2, 3, figsize=(20, 12))
fig6.suptitle(f'PnL Distribution (annualised) — Budget ${ANNUAL_BUDGET_USD/1e6:.0f}M/yr',
              fontsize=14, fontweight='bold')

for idx, money in enumerate(focus_moneys):
    ax = axes2[idx // 3, idx % 3]
    sub = df_agg[df_agg['moneyness'] == money]
    mats_avail = sorted(sub['maturity'].unique())

    box_data = []
    box_labels = []
    for mat in mats_avail:
        vals = sub[sub['maturity'] == mat]['pnl_annualized_pct'].values
        if len(vals) > 0:
            box_data.append(vals)
            box_labels.append(f'{mat}j')

    if box_data:
        bp = ax.boxplot(box_data, labels=box_labels, patch_artist=True,
                        showfliers=True, flierprops=dict(markersize=3, alpha=0.3))
        for patch in bp['boxes']:
            patch.set_facecolor(colors_money.get(money, 'gray'))
            patch.set_alpha(0.5)

    ax.axhline(y=0, color='black', linewidth=0.8)
    ax.axhline(y=-ANNUAL_BUDGET_PCT_DISPLAY, color=COL['ref'], linewidth=0.5, linestyle='--', alpha=0.5)
    ax.set_title(f'Moneyness = {money}%', fontweight='bold')
    ax.set_ylabel('PnL (annualised - %)')
    ax.set_xlabel('Maturity')
    ax.grid(axis='y', alpha=0.3)

plt.tight_layout()
plt.savefig(f'{folder_name}/boxplots_pnl.png', dpi=150, bbox_inches='tight')
print("Plot 6: boxplots_pnl.png")

# %%
# ── Plot 7: SCR Before vs After ──
colors_inv = ["#0D47A1", "#1E88E5", "#90CAF9", "#E0E0E0"]
cmap_scr_inv = LinearSegmentedColormap.from_list("BlueToGray", colors_inv)

pivot_nu = dist['SCR_nu_mean'].unstack()
pivot_hedged = dist['SCR_hedged_avg'].unstack()

fig7, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 8.5), sharey=True)

vmin = min(pivot_hedged.min().min(), pivot_nu.min().min())
vmax = max(pivot_hedged.max().max(), pivot_nu.max().max())

im1 = ax1.imshow(pivot_nu.values, cmap=cmap_scr_inv, aspect='auto', vmin=vmin, vmax=vmax)
ax1.set_title('Average SCR Before Hedges (%)\n(Initial High Risk)', fontsize=13, fontweight='bold', pad=15)
ax1.set_ylabel('Moneyness (Strike %)', fontsize=11)

im2 = ax2.imshow(pivot_hedged.values, cmap=cmap_scr_inv, aspect='auto', vmin=vmin, vmax=vmax)
ax2.set_title('Average SCR After Overlay (%)\n(Hedged Capital Efficiency)', fontsize=13, fontweight='bold', pad=15)

for ax, data in zip([ax1, ax2], [pivot_nu, pivot_hedged]):
    ax.set_xticks(range(len(data.columns)))
    ax.set_xticklabels([f'{int(c)}j' for c in data.columns], fontsize=9)
    ax.set_yticks(range(len(data.index)))
    ax.set_yticklabels([f'K={int(m)}%' for m in data.index], fontsize=9)
    ax.set_xlabel('Maturity (days)', fontsize=11)

    for i in range(len(data.index)):
        for j in range(len(data.columns)):
            val = data.values[i, j]
            if not np.isnan(val):
                txt_color = "white" if val < (vmin + (vmax - vmin) * 0.35) else "black"
                ax.text(j, i, f'{val:.1f}%', ha='center', va='center',
                        fontsize=9, fontweight='bold', color=txt_color)

cbar_ax = fig7.add_axes([0.93, 0.15, 0.015, 0.7])
fig7.colorbar(im1, cax=cbar_ax, label='SCR level (%) - Blue is Better')

fig7.suptitle(f'Prudential Impact: Blue = Capital Relief\n'
              f'Budget: ${ANNUAL_BUDGET_USD/1e6:.0f}M p.a. | {START_DATE} to {END_DATE}',
              fontsize=15, fontweight='bold', y=1.02)

plt.tight_layout(rect=[0, 0, 0.92, 1])
plt.savefig(f"{folder_name}/scr_comparison_inverted_blue.png", dpi=150, bbox_inches='tight')
plt.show()
print("Plot 7: scr_comparison_inverted_blue.png")

# %%
# ── Plot 8: Symmetrical Adjustment ──
fig8, axb = plt.subplots(figsize=(14, 6))

shock_ts = 0.39 + sa_clipped

axb.plot(shock_ts.index, shock_ts.values * 100, color='darkred', linewidth=1.8, label='Total Equity Shock (39% + SA)')
axb.axhline(y=39, color='black', linewidth=1.2, linestyle='-', alpha=0.6, label='Base Charge (39%)')
axb.axhline(y=49, color='#C0392B', linewidth=1, linestyle='--', label='Regulatory Cap (49%)')
axb.axhline(y=29, color='#27AE60', linewidth=1, linestyle='--', label='Regulatory Floor (29%)')

axb.fill_between(shock_ts.index, 39, shock_ts.values * 100,
                 where=(shock_ts.values * 100 >= 39), color='#C0392B', alpha=0.15)
axb.fill_between(shock_ts.index, 39, shock_ts.values * 100,
                 where=(shock_ts.values * 100 < 39), color='#27AE60', alpha=0.15)

axb.set_ylabel('Capital Charge (SCR Equity %)', fontsize=11, fontweight='bold')
axb.set_title('Solvency II Symmetrical Adjustment (SA) Evolution\n'
              'Standard Formula: Equity Type 1 (Global)', fontsize=14, fontweight='bold', pad=15)
axb.set_ylim(25, 53)
axb.grid(True, linestyle=':', alpha=0.5)
axb.legend(loc='upper left', fontsize=10, ncol=2, frameon=True)
axb.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
axb.xaxis.set_major_locator(mdates.YearLocator())

plt.tight_layout()
plt.savefig(f"{folder_name}/solvency_ii_sa_shock.png", dpi=150, bbox_inches='tight')
plt.show()
print("Plot 8: solvency_ii_sa_shock.png")

# %%
# ── Plot 9: Boxplots M net ──
fig9, axes3 = plt.subplots(2, 3, figsize=(20, 12))
fig9.suptitle(f'Net Capital Benefit (M) Distribution — {START_DATE} - {END_DATE}',
              fontsize=14, fontweight='bold')

for idx, money in enumerate(focus_moneys):
    ax = axes3[idx // 3, idx % 3]
    sub = df_agg[df_agg['moneyness'] == money]
    mats_avail = sorted(sub['maturity'].unique())

    box_data = []
    box_labels = []
    for mat in mats_avail:
        vals = sub[sub['maturity'] == mat]['M_mean'].values
        if len(vals) > 0:
            box_data.append(vals)
            box_labels.append(f'{mat}j')

    if box_data:
        bp = ax.boxplot(box_data, labels=box_labels, patch_artist=True,
                        showfliers=True, flierprops=dict(markersize=3, alpha=0.3))
        for patch in bp['boxes']:
            patch.set_facecolor(colors_money.get(money, 'gray'))
            patch.set_alpha(0.5)

    ax.axhline(y=0, color='black', linewidth=0.8, linestyle='--')
    ax.set_title(f'Moneyness = {money}%', fontweight='bold')
    ax.set_ylabel('Avg Net Capital Benefit M (%)')
    ax.set_xlabel('Maturity')
    ax.grid(axis='y', alpha=0.3)

plt.tight_layout()
plt.savefig(f'{folder_name}/boxplots_M_net.png', dpi=150, bbox_inches='tight')
print("Plot 9: boxplots_M_net.png")

# %%
# ═══════════════════════════════════════════════════════════════════════
# 7. EXPORT — CSV + PDF + EXCEL DETAILLE
# ═══════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("7. EXPORT")
print("=" * 70)

# PDF rapport
with PdfPages(f'{folder_name}/Rapport_Rolling_Entry.pdf') as pdf:
    for fig_obj in [fig0, fig1, fig2, fig3, fig4, fig5, fig6, fig7, fig8, fig9]:
        pdf.savefig(fig_obj, bbox_inches='tight')
    d = pdf.infodict()
    d['Title'] = f'Rolling Backtest — Puts S&P 500 - {START_DATE} - {END_DATE}'

# CSV
dist.to_csv(f'{folder_name}/distributional_summary.csv')
df_agg.to_csv(f'{folder_name}/all_backtests_agg.csv', index=False)

# ═══════════════════════════════════════════════════════════════════════
# EXCEL DETAILLE — Pour verification de consistance
# ═══════════════════════════════════════════════════════════════════════

print("\nGeneration de l'Excel detaille...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─────────────────────────────────────────────────
# SHEET 1: PARAMETRES
# ─────────────────────────────────────────────────
ws_params = wb.active
ws_params.title = "Parameters"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1B2A4A")
blue_font = Font(color="0000FF", bold=False, size=10)
data_font = Font(size=10)
pct_format = '0.00%'
usd_format = '#,##0'
usd_dec_format = '#,##0.00'
num_format = '#,##0.0000'

params = [
    ("Parameter", "Value", "Unit"),
    ("Annual Budget", ANNUAL_BUDGET_USD, "USD"),
    ("MV Equity (t0)", MV_EQUITY_T0, "USD"),
    ("Spot (t0)", SPOT_T0, "USD"),
    ("Budget / MV_t0", ANNUAL_BUDGET_USD / MV_EQUITY_T0, "%"),
    ("Start Date", START_DATE, ""),
    ("End Date", END_DATE, ""),
    ("Window", WINDOW_YEARS, "years"),
    ("Window", WINDOW_DAYS, "days"),
    ("N Starts", N_STARTS_ACTUAL, ""),
    ("N Instruments", df_agg.groupby(['moneyness','maturity']).ngroups, ""),
    ("Total Detail Records", len(df_detail), ""),
    ("IV Mean (global)", iv_avg_global, ""),
]

for r, row_data in enumerate(params, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws_params.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        elif c == 2:
            cell.font = blue_font
            if isinstance(val, (int, float)) and val > 1000:
                cell.number_format = usd_format

ws_params.column_dimensions['A'].width = 25
ws_params.column_dimensions['B'].width = 20
ws_params.column_dimensions['C'].width = 12

# ─────────────────────────────────────────────────
# SHEET 2: DISTRIBUTIONAL SUMMARY
# ─────────────────────────────────────────────────
ws_dist = wb.create_sheet("Distributional Summary")

dist_reset = dist.reset_index()
headers_dist = list(dist_reset.columns)

for c, h in enumerate(headers_dist, 1):
    cell = ws_dist.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for r, (_, row_data) in enumerate(dist_reset.iterrows(), 2):
    for c, h in enumerate(headers_dist, 1):
        val = row_data[h]
        if pd.isna(val):
            val = ""
        cell = ws_dist.cell(row=r, column=c, value=val)
        cell.font = data_font
        if isinstance(val, float):
            cell.number_format = '0.0000'

for c in range(1, len(headers_dist) + 1):
    ws_dist.column_dimensions[get_column_letter(c)].width = 18

# ─────────────────────────────────────────────────
# SHEET 3: DETAIL ROLLS (premier backtest start_id=0)
# Ceci permet de verifier ligne par ligne le pricing
# ─────────────────────────────────────────────────
ws_detail = wb.create_sheet("Detail Rolls (Start 0)")

df_start0 = df_detail[df_detail['start_id'] == 0].copy()
df_start0 = df_start0.sort_values(['moneyness', 'maturity', 'date_buy']).reset_index(drop=True)

# Colonnes selectionnees pour lisibilite
detail_cols = [
    'moneyness', 'maturity', 'date_buy', 'date_exp', 'tau_remaining', 'is_mtm_exit',
    'S_buy', 'S_exp', 'F_buy', 'K', 'iv_buy', 'r_buy', 'q_buy',
    'return_underlying',
    'MV_equity',
    'put_price_per_unit', 'put_price_pct_spot', 'put_shocked_per_unit', 'hedge_gain_per_put',
    'budget_per_roll_usd', 'nb_puts', 'premium_paid_usd', 'transaction_fees_usd', 'total_cost_usd',
    'premium_pct_mv', 'fees_pct_premium',
    'equity_shock', 'SCR_nu_usd', 'SCR_hedged_usd',
    'gross_scr_saving_usd', 'net_capital_benefit_usd',
    'SCR_nu_pct', 'SCR_hedged_pct', 'gross_scr_saving_pct', 'M_net_pct', 'efficiency_ratio',
    'payoff_per_put', 'payoff_usd', 'pnl_roll_usd', 'pnl_roll_pct',
]

# Filtrer les colonnes disponibles
detail_cols = [c for c in detail_cols if c in df_start0.columns]

# Headers
for c, h in enumerate(detail_cols, 1):
    cell = ws_detail.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Data
for r, (_, row_data) in enumerate(df_start0.iterrows(), 2):
    for c, h in enumerate(detail_cols, 1):
        val = row_data[h]
        if isinstance(val, pd.Timestamp):
            val = val.strftime('%Y-%m-%d')
        elif pd.isna(val):
            val = ""
        cell = ws_detail.cell(row=r, column=c, value=val)
        cell.font = data_font
        if isinstance(val, float):
            if 'usd' in h.lower() or h in ['budget_per_roll_usd', 'premium_paid_usd', 'total_cost_usd',
                                             'SCR_nu_usd', 'SCR_hedged_usd', 'gross_scr_saving_usd',
                                             'net_capital_benefit_usd', 'payoff_usd', 'pnl_roll_usd',
                                             'MV_equity']:
                cell.number_format = '#,##0'
            elif 'pct' in h.lower() or h in ['iv_buy', 'r_buy', 'q_buy', 'equity_shock',
                                              'return_underlying', 'pnl_roll_pct', 'M_net_pct',
                                              'gross_scr_saving_pct', 'SCR_nu_pct', 'SCR_hedged_pct',
                                              'premium_pct_mv', 'fees_pct_premium']:
                cell.number_format = '0.0000'
            else:
                cell.number_format = '0.0000'

for c in range(1, len(detail_cols) + 1):
    ws_detail.column_dimensions[get_column_letter(c)].width = 18

# Freeze header
ws_detail.freeze_panes = 'A2'
ws_dist.freeze_panes = 'A2'

# ─────────────────────────────────────────────────
# SHEET 4: ALL BACKTESTS AGGREGATED
# ─────────────────────────────────────────────────
ws_agg = wb.create_sheet("All Backtests Agg")

agg_cols = list(df_agg.columns)

for c, h in enumerate(agg_cols, 1):
    cell = ws_agg.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for r, (_, row_data) in enumerate(df_agg.iterrows(), 2):
    for c, h in enumerate(agg_cols, 1):
        val = row_data[h]
        if isinstance(val, pd.Timestamp):
            val = val.strftime('%Y-%m-%d')
        elif pd.isna(val):
            val = ""
        cell = ws_agg.cell(row=r, column=c, value=val)
        cell.font = data_font
        if isinstance(val, float):
            if 'usd' in h.lower():
                cell.number_format = '#,##0'
            else:
                cell.number_format = '0.0000'

for c in range(1, len(agg_cols) + 1):
    ws_agg.column_dimensions[get_column_letter(c)].width = 18

ws_agg.freeze_panes = 'A2'

# ─────────────────────────────────────────────────
# SHEET 5: SAMPLE VERIFICATION (quelques rolls pour audit rapide)
# ─────────────────────────────────────────────────
ws_verif = wb.create_sheet("Verification Sample")

# Prendre 1 exemple par instrument (premier roll du start 0)
sample = df_start0.groupby(['moneyness', 'maturity']).first().reset_index()

verif_cols = [
    'moneyness', 'maturity', 'date_buy', 'S_buy', 'F_buy', 'K', 'iv_buy',
    'put_price_per_unit', 'put_price_pct_spot',
    'budget_per_roll_usd', 'nb_puts', 'premium_paid_usd', 'transaction_fees_usd',
    'MV_equity', 'equity_shock', 'SCR_nu_usd',
    'put_shocked_per_unit', 'hedge_gain_per_put',
    'gross_scr_saving_usd', 'net_capital_benefit_usd', 'M_net_pct', 'efficiency_ratio',
    'S_exp', 'payoff_per_put', 'payoff_usd', 'pnl_roll_usd', 'pnl_roll_pct',
]
verif_cols = [c for c in verif_cols if c in sample.columns]

# Title row
ws_verif.cell(row=1, column=1, value="VERIFICATION SAMPLE — 1 roll per instrument (Start 0)")
ws_verif.cell(row=1, column=1).font = Font(bold=True, size=12)
ws_verif.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(verif_cols))

# Formulas legend
legend = [
    ("Key Formulas:", ""),
    ("F_buy", "= S_buy * exp((r - q) * tau)"),
    ("K", "= (moneyness/100) * F_buy"),
    ("put_price", "= BS_Put(S_buy, K, tau, r, q, iv)"),
    ("nb_puts", "= budget_per_roll_usd / unit_cost_all_in"),
    ("premium_paid", "= nb_puts * put_price"),
    ("MV_equity", "= MV_t0 * (S_buy / S_t0)"),
    ("SCR_nu_usd", "= MV_equity * equity_shock"),
    ("gross_scr_saving", "= SCR_nu - max(0, SCR_nu - nb_puts*(put_shocked - put_price))"),
    ("net_capital_benefit", "= gross_scr_saving - premium_paid"),
    ("M_net_pct", "= net_capital_benefit / MV_equity * 100"),
    ("pnl_roll_usd", "= payoff_usd - budget_per_roll_usd"),
    ("pnl_roll_pct", "= pnl_roll_usd / MV_equity * 100"),
]

for r, (name, formula) in enumerate(legend, 3):
    ws_verif.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
    ws_verif.cell(row=r, column=2, value=formula).font = Font(italic=True, size=10, color="0000FF")

data_start_row = 3 + len(legend) + 1

for c, h in enumerate(verif_cols, 1):
    cell = ws_verif.cell(row=data_start_row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for r, (_, row_data) in enumerate(sample.iterrows(), data_start_row + 1):
    for c, h in enumerate(verif_cols, 1):
        val = row_data[h]
        if isinstance(val, pd.Timestamp):
            val = val.strftime('%Y-%m-%d')
        elif pd.isna(val):
            val = ""
        cell = ws_verif.cell(row=r, column=c, value=val)
        cell.font = data_font
        if isinstance(val, float):
            if 'usd' in h.lower() or h in ['MV_equity', 'budget_per_roll_usd', 'premium_paid_usd',
                                             'transaction_fees_usd', 'SCR_nu_usd', 'gross_scr_saving_usd',
                                             'net_capital_benefit_usd', 'payoff_usd', 'pnl_roll_usd']:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '0.0000'

for c in range(1, len(verif_cols) + 1):
    ws_verif.column_dimensions[get_column_letter(c)].width = 20

ws_verif.freeze_panes = f'A{data_start_row + 1}'

# Save
excel_path = f'{folder_name}/Backtest_Detail_USD.xlsx'
wb.save(excel_path)
print(f"Excel detaille: {excel_path}")
print(f"  - Sheet 'Parameters': parametres du backtest")
print(f"  - Sheet 'Distributional Summary': resume par instrument")
print(f"  - Sheet 'Detail Rolls (Start 0)': {len(df_start0)} rolls detailles")
print(f"  - Sheet 'All Backtests Agg': {len(df_agg)} records agreges")
print(f"  - Sheet 'Verification Sample': 1 roll/instrument + formules")

print(f"\nRapport: {folder_name}/Rapport_Rolling_Entry.pdf")
print(f"Summary: {folder_name}/distributional_summary.csv")
print(f"Detail:  {folder_name}/all_backtests_agg.csv")

print("\n" + "=" * 70)
print("TERMINE")
print("=" * 70)

# %%
